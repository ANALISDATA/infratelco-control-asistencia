"""Generación del Excel corporativo de asistencia (openpyxl). Ver documentation/excel.md
-- esto cubre el reporte de "Asistencia del día"; el Excel completo multi-hoja
(Resumen Gerencial, Empleados, Indicadores) queda para la Fase 4.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend import config

_AZUL_OSCURO = "00226E"
_BLANCO = "FFFFFF"
_GRIS_TEXTO = "6B7280"
_GRIS_FILA = "F4F6FA"
_BORDE_COLOR = "D9DCE3"

_LOGO_PATH = config.RAIZ / "assets" / "logos" / "infratelco_logo.png"

_FILA_TITULO = 1
_FILA_SUBTITULO = 2
_FILA_ENCABEZADOS = 4
_FILA_DATOS_INICIO = 5


def generar_reporte_asistencia(subtitulo: str, filas: list[dict]) -> bytes:
    """`subtitulo`: texto libre bajo el título (ej. "Fecha: 30/08/2026" o "Del
    16/08/2026 al 30/08/2026"). `filas`: lista de dicts con las mismas claves (mismo
    orden = columnas del Excel), tal como ya se arman para la tabla en pantalla.
    Devuelve los bytes del .xlsx, listos para `st.download_button`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    columnas = list(filas[0].keys()) if filas else ["Sin registros"]
    ultima_col_letra = get_column_letter(len(columnas))

    if _LOGO_PATH.exists():
        imagen = XLImage(str(_LOGO_PATH))
        alto_objetivo = 78
        escala = alto_objetivo / imagen.height
        imagen.height = alto_objetivo
        imagen.width = int(imagen.width * escala)
        ws.add_image(imagen, "A1")
        col_inicio = 3  # deja A y B libres para que no tape el logo
    else:
        col_inicio = 1
    col_inicio_letra = get_column_letter(col_inicio)

    ws.merge_cells(f"{col_inicio_letra}{_FILA_TITULO}:{ultima_col_letra}{_FILA_TITULO}")
    celda_titulo = ws[f"{col_inicio_letra}{_FILA_TITULO}"]
    celda_titulo.value = "INFRATELCO — Detalle de registros ingresos y salidas"
    celda_titulo.font = Font(bold=True, size=15, color=_AZUL_OSCURO)
    celda_titulo.alignment = Alignment(vertical="center")

    ws.merge_cells(f"{col_inicio_letra}{_FILA_SUBTITULO}:{ultima_col_letra}{_FILA_SUBTITULO}")
    celda_subtitulo = ws[f"{col_inicio_letra}{_FILA_SUBTITULO}"]
    celda_subtitulo.value = subtitulo
    celda_subtitulo.font = Font(size=11, color=_GRIS_TEXTO)
    celda_subtitulo.alignment = Alignment(vertical="center")

    ws.row_dimensions[_FILA_TITULO].height = 28
    ws.row_dimensions[_FILA_SUBTITULO].height = 20

    borde = Border(*[Side(style="thin", color=_BORDE_COLOR)] * 4)

    for i, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=_FILA_ENCABEZADOS, column=i, value=nombre)
        celda.font = Font(bold=True, color=_BLANCO, size=11)
        celda.fill = PatternFill("solid", fgColor=_AZUL_OSCURO)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde
    ws.row_dimensions[_FILA_ENCABEZADOS].height = 24

    for offset, fila in enumerate(filas):
        fila_idx = _FILA_DATOS_INICIO + offset
        for col_idx, nombre in enumerate(columnas, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=fila.get(nombre))
            celda.border = borde
            celda.alignment = Alignment(vertical="center")
            if offset % 2 == 1:
                celda.fill = PatternFill("solid", fgColor=_GRIS_FILA)

    for i, nombre in enumerate(columnas, start=1):
        ancho_contenido = max((len(str(f.get(nombre, ""))) for f in filas), default=0)
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(nombre), ancho_contenido) + 3, 40)

    ws.freeze_panes = f"A{_FILA_DATOS_INICIO}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
